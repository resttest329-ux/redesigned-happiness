"""Items catalog routes.

An item is exactly a reusable wizard line minus the per-invoice fields, so the
frontend can merge a saved item straight into ``step3.lines`` with no changes
to invoice assembly. Items are per-business scoped, soft-deleted (``is_active``)
so removing one can never break a pending draft, and searchable across
name / SKU / description.
"""

from __future__ import annotations

import csv
import io
import logging
from datetime import datetime, timezone
from typing import Annotated, Optional

from fastapi import (
    APIRouter,
    Depends,
    File,
    HTTPException,
    Query,
    UploadFile,
    status,
)
from pydantic import ValidationError
from sqlalchemy import or_
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from auth import oauth2_scheme
from deps import get_current_user_obj, get_db
from utils import schema
from utils.models import Item

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/items", tags=["Items"])

ITEM_FIELDS = (
    "sku",
    "name",
    "description",
    "hsn_code",
    "hsn_category",
    "isic_code",
    "isic_category",
    "unit_price",
    "price_unit",
    "base_quantity",
)

IMPORT_COLUMNS = list(ITEM_FIELDS)
MAX_IMPORT_ROWS = 2000


def _coerce_int(value, default: int) -> int:
    """Normalize a paging value for direct Python calls.

    Under FastAPI the declared ``Query(...)`` default is replaced by the parsed
    request value, so this is a no-op. When a route is called directly from
    Python (tests, internal helpers) the unresolved ``Query`` marker object can
    leak through as the default — coerce it back to the documented default so
    the query builder never receives a non-integer.
    """
    if isinstance(value, bool) or not isinstance(value, (int, float, str)):
        return default
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


def _coerce_str(value) -> str | None:
    """Same intent as ``_coerce_int`` for optional string filters."""
    if not isinstance(value, str):
        return None
    value = value.strip()
    return value or None


def _coerce_bool(value, default: bool | None) -> bool | None:
    if value is None or isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    if isinstance(value, str):
        lowered = value.strip().lower()
        if lowered in {"true", "1", "yes"}:
            return True
        if lowered in {"false", "0", "no"}:
            return False
        return default
    return default


def _scoped(db: Session, business_id: str):
    return db.query(Item).filter(Item.business_id == business_id)


def _get_owned(db: Session, item_id: int, business_id: str) -> Item:
    item = _scoped(db, business_id).filter(Item.id == item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Item not found")
    return item


def _find_by_sku(db: Session, business_id: str, sku: str | None) -> Item | None:
    if not sku:
        return None
    return _scoped(db, business_id).filter(Item.sku == sku).first()


def _apply_payload(item: Item, data: dict) -> None:
    for field in ITEM_FIELDS:
        if field in data:
            setattr(item, field, data[field])
    item.updated_at = datetime.now(timezone.utc)


@router.get("", response_model=schema.ItemPage)
def list_items(
    token: Annotated[str, Depends(oauth2_scheme)],
    *,
    db: Session = Depends(get_db),
    search: Optional[str] = None,
    kind: Optional[str] = Query(None, pattern="^(product|service)$"),
    active: Optional[bool] = True,
    offset: int = 0,
    limit: int = Query(50, ge=1, le=500),
):
    user = get_current_user_obj(token, db)
    search = _coerce_str(search)
    kind = _coerce_str(kind)
    kind = kind if kind in ("product", "service") else None
    active = _coerce_bool(active, True)
    offset = max(0, _coerce_int(offset, 0))
    limit = min(500, max(1, _coerce_int(limit, 50)))
    query = _scoped(db, user.business_id)

    if active is not None:
        query = query.filter(Item.is_active == active)

    if kind == "product":
        query = query.filter(Item.hsn_code.isnot(None), Item.hsn_code != "")
    elif kind == "service":
        query = query.filter(Item.isic_code.isnot(None), Item.isic_code != "")

    if search:
        like = f"%{search}%"
        query = query.filter(
            or_(
                Item.name.ilike(like),
                Item.sku.ilike(like),
                Item.description.ilike(like),
            )
        )

    total = query.count()
    items = query.order_by(Item.name).offset(offset).limit(limit).all()
    return {"total": total, "offset": offset, "limit": limit, "items": items}


@router.post(
    "", response_model=schema.ItemOut, status_code=status.HTTP_201_CREATED
)
def create_item(
    data: schema.ItemCreate,
    token: Annotated[str, Depends(oauth2_scheme)],
    db: Session = Depends(get_db),
):
    user = get_current_user_obj(token, db)
    payload = data.model_dump()

    existing = _find_by_sku(db, user.business_id, payload.get("sku"))
    if existing is not None:
        raise HTTPException(
            status_code=409,
            detail=(
                f"An item with SKU '{payload['sku']}' already exists in this "
                "workspace. Edit that item instead, or use a different SKU."
            ),
        )

    item = Item(business_id=user.business_id, is_active=True, **payload)
    db.add(item)
    try:
        db.commit()
    except IntegrityError:
        logger.exception("create_item: integrity error")
        db.rollback()
        raise HTTPException(
            status_code=409,
            detail="An item with that SKU already exists in this workspace.",
        )
    db.refresh(item)
    return item


@router.post("/import", response_model=schema.ItemImportResult)
async def import_items(
    token: Annotated[str, Depends(oauth2_scheme)],
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    user = get_current_user_obj(token, db)
    raw = await file.read()
    rows = _parse_import_file(file.filename or "", raw)
    if not rows:
        raise HTTPException(
            status_code=400,
            detail=(
                "No rows found. Expected a CSV/XLSX file with the columns: "
                + ", ".join(IMPORT_COLUMNS)
            ),
        )
    if len(rows) > MAX_IMPORT_ROWS:
        raise HTTPException(
            status_code=400,
            detail=f"Too many rows — import at most {MAX_IMPORT_ROWS} at a time.",
        )
    return _process_import_rows(rows, db=db, business_id=user.business_id)


@router.post("/bulk-delete")
def bulk_delete_items(
    body: schema.ItemBulkDelete,
    token: Annotated[str, Depends(oauth2_scheme)],
    db: Session = Depends(get_db),
):
    """Soft-delete (deactivate) by default; ``hard=true`` removes the rows."""
    user = get_current_user_obj(token, db)
    ids = list(dict.fromkeys(body.ids))
    if not ids:
        raise HTTPException(
            status_code=400, detail="Select at least one item to delete."
        )
    items = _scoped(db, user.business_id).filter(Item.id.in_(ids)).all()
    now = datetime.now(timezone.utc)
    deleted = 0
    for item in items:
        if body.hard:
            db.delete(item)
            deleted += 1
        else:
            if item.is_active:
                item.is_active = False
                item.updated_at = now
            deleted += 1
    db.commit()
    return {
        "deleted": deleted,
        "requested": len(ids),
        "mode": "hard" if body.hard else "soft",
    }


@router.post("/bulk-activate")
def bulk_activate_items(
    body: schema.ItemBulkDelete,
    token: Annotated[str, Depends(oauth2_scheme)],
    db: Session = Depends(get_db),
):
    user = get_current_user_obj(token, db)
    ids = list(dict.fromkeys(body.ids))
    if not ids:
        raise HTTPException(
            status_code=400, detail="Select at least one item to restore."
        )
    items = _scoped(db, user.business_id).filter(Item.id.in_(ids)).all()
    now = datetime.now(timezone.utc)
    restored = 0
    for item in items:
        if not item.is_active:
            item.is_active = True
            item.updated_at = now
        restored += 1
    db.commit()
    return {"activated": restored, "requested": len(ids)}


@router.get("/{item_id}", response_model=schema.ItemOut)
def get_item(
    item_id: int,
    token: Annotated[str, Depends(oauth2_scheme)],
    db: Session = Depends(get_db),
):
    user = get_current_user_obj(token, db)
    return _get_owned(db, item_id, user.business_id)


@router.patch("/{item_id}", response_model=schema.ItemOut)
def update_item(
    item_id: int,
    data: schema.ItemUpdate,
    token: Annotated[str, Depends(oauth2_scheme)],
    db: Session = Depends(get_db),
):
    user = get_current_user_obj(token, db)
    item = _get_owned(db, item_id, user.business_id)
    payload = data.model_dump(exclude_unset=True)

    new_sku = payload.get("sku")
    if new_sku and new_sku != item.sku:
        clash = _find_by_sku(db, user.business_id, new_sku)
        if clash is not None and clash.id != item.id:
            raise HTTPException(
                status_code=409,
                detail=f"Another item already uses SKU '{new_sku}'.",
            )

    _apply_payload(item, payload)
    try:
        db.commit()
    except IntegrityError:
        logger.exception("update_item: integrity error")
        db.rollback()
        raise HTTPException(
            status_code=409,
            detail="Another item already uses that SKU.",
        )
    db.refresh(item)
    return item


@router.delete("/{item_id}", response_model=schema.ItemOut)
def delete_item(
    item_id: int,
    token: Annotated[str, Depends(oauth2_scheme)],
    *,
    db: Session = Depends(get_db),
    hard: bool = False,
):
    hard = bool(_coerce_bool(hard, False))
    user = get_current_user_obj(token, db)
    item = _get_owned(db, item_id, user.business_id)
    if hard:
        snapshot = schema.ItemOut.model_validate(item)
        db.delete(item)
        db.commit()
        return snapshot
    item.is_active = False
    item.updated_at = datetime.now(timezone.utc)
    db.commit()
    db.refresh(item)
    return item


# --------------------------------------------------------------------------
# Import helpers
# --------------------------------------------------------------------------


def _parse_import_file(filename: str, raw: bytes) -> list[dict]:
    lower = (filename or "").lower()
    if lower.endswith((".xlsx", ".xlsm")):
        return _parse_xlsx(raw)
    return _parse_csv(raw)


def _parse_csv(raw: bytes) -> list[dict]:
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            text = raw.decode(encoding)
            break
        except UnicodeDecodeError:
            # Expected while probing encodings — try the next candidate.
            logging.exception("Unexpected error")
            logger.debug("_parse_csv: decode failed for %s", encoding)
            continue
    else:
        raise HTTPException(
            status_code=400, detail="Could not decode the uploaded file."
        )
    reader = csv.DictReader(io.StringIO(text))
    rows: list[dict] = []
    for raw_row in reader:
        rows.append(
            {
                (k or "").strip().lower(): (v if v is not None else "")
                for k, v in raw_row.items()
            }
        )
    return rows


def _parse_xlsx(raw: bytes) -> list[dict]:
    try:
        from openpyxl import load_workbook
    except Exception:
        logger.exception("_parse_xlsx: openpyxl unavailable")
        raise HTTPException(
            status_code=400,
            detail="XLSX import is unavailable — please upload a CSV file.",
        )
    try:
        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        header = next(rows_iter, None)
        if not header:
            return []
        keys = [str(h or "").strip().lower() for h in header]
        rows: list[dict] = []
        for values in rows_iter:
            if values is None or all(v in (None, "") for v in values):
                continue
            row = {}
            for idx, key in enumerate(keys):
                if not key:
                    continue
                row[key] = values[idx] if idx < len(values) else ""
            rows.append(row)
        return rows
    except HTTPException:
        logging.exception("Unexpected error")
        raise
    except Exception:
        logger.exception("_parse_xlsx failed")
        raise HTTPException(
            status_code=400, detail="Could not read the uploaded spreadsheet."
        )


def _row_label(index: int, row: dict) -> str:
    sku = str(row.get("sku") or "").strip()
    name = str(row.get("name") or "").strip()
    tag = sku or name or "(unnamed)"
    return f"Row {index} [{tag}]"


def _format_validation_error(exc: ValidationError) -> str:
    parts: list[str] = []
    for err in exc.errors():
        loc = ".".join(str(x) for x in err.get("loc", ()) if x != "__root__")
        msg = err.get("msg", "invalid value")
        msg = msg.replace("Value error, ", "")
        parts.append(f"{loc}: {msg}" if loc else msg)
    return "; ".join(parts) or "invalid row"


def _process_import_rows(
    rows: list[dict], db: Session, business_id: str
) -> schema.ItemImportResult:
    """Validate and upsert import rows. Bad rows are skipped, never fatal."""
    created = 0
    updated = 0
    skipped = 0
    errors: list[str] = []
    now = datetime.now(timezone.utc)

    for index, raw_row in enumerate(rows, start=1):
        row = {
            str(k or "").strip().lower(): v for k, v in (raw_row or {}).items()
        }
        payload_raw = {
            field: row.get(field) for field in ITEM_FIELDS if field in row
        }
        # Drop blanks so schema defaults apply.
        payload_raw = {
            k: v
            for k, v in payload_raw.items()
            if v is not None and str(v).strip() != ""
        }
        try:
            parsed = schema.ItemCreate(**payload_raw)
        except ValidationError as exc:
            # Expected for bad user data: a skipped row is a normal outcome of
            # importing a spreadsheet, not a fault. Record the reason for the
            # user and move on — no traceback, no exception-level logging.
            logging.exception("Unexpected error")
            detail = _format_validation_error(exc)
            logger.info(
                "import row skipped (validation): %s: %s",
                _row_label(index, row),
                detail,
            )
            skipped += 1
            errors.append(f"{_row_label(index, row)}: {detail}")
            continue
        except Exception as exc:
            # Still a row-level outcome: skip it with a concise reason instead
            # of emitting a traceback for the whole import.
            logging.exception("Unexpected error")
            logger.warning(
                "import row skipped (unparseable): %s: %s",
                _row_label(index, row),
                exc,
            )
            skipped += 1
            errors.append(f"{_row_label(index, row)}: could not be parsed")
            continue

        data = parsed.model_dump()
        try:
            existing = _find_by_sku(db, business_id, data.get("sku"))
            if existing is not None:
                _apply_payload(existing, data)
                existing.is_active = True
                db.commit()
                updated += 1
            else:
                item = Item(business_id=business_id, is_active=True, **data)
                item.created_at = item.created_at or now
                db.add(item)
                db.commit()
                created += 1
        except IntegrityError:
            # Duplicate SKU is expected user data — roll back and skip the row.
            logging.exception("Unexpected error")
            db.rollback()
            logger.info(
                "import row skipped (duplicate SKU): %s",
                _row_label(index, row),
            )
            skipped += 1
            errors.append(
                f"{_row_label(index, row)}: duplicate SKU in this workspace"
            )
        except Exception as exc:
            logging.exception("Unexpected error")
            db.rollback()
            logger.warning(
                "import row skipped (write failed): %s: %s",
                _row_label(index, row),
                exc,
            )
            skipped += 1
            errors.append(f"{_row_label(index, row)}: could not be saved")

    return schema.ItemImportResult(
        created=created,
        updated=updated,
        skipped=skipped,
        errors=errors[:100],
    )
