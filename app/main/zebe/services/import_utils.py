"""Shared spreadsheet import helpers.

Both the items catalog and the customer directory accept the same style of
bulk upload: a CSV (or XLSX when ``openpyxl`` is installed) whose header row
names the payload fields. Parsing, row labelling and validation-error
formatting are identical for both, so they live here once.

Row-level failures are always a normal outcome of importing a user
spreadsheet, never a fault of the service: callers skip the row, record a
concise reason and carry on.
"""

from __future__ import annotations

import csv
import io
import logging

from fastapi import HTTPException
from pydantic import ValidationError

logger = logging.getLogger(__name__)

MAX_IMPORT_ROWS = 2000


def parse_import_file(filename: str, raw: bytes) -> list[dict]:
    """Return the uploaded sheet as a list of lower-cased key dicts."""
    lower = (filename or "").lower()
    if lower.endswith((".xlsx", ".xlsm")):
        return _parse_xlsx(raw)
    return _parse_csv(raw)


def _parse_csv(raw: bytes) -> list[dict]:
    text = None
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            text = raw.decode(encoding)
            break
        except UnicodeDecodeError:
            # Expected while probing encodings: try the next candidate and
            # keep the log concise (no traceback).
            logging.exception("Unexpected error")
            logger.debug("parse_csv: decode failed for %s", encoding)
            continue
    if text is None:
        raise HTTPException(
            status_code=400, detail="Could not decode the uploaded file."
        )
    reader = csv.DictReader(io.StringIO(text))
    rows: list[dict] = []
    for raw_row in reader:
        rows.append(
            {
                (k or "").strip().lower(): (v if v is not None else "")
                for k, v in raw_row.items()
            }
        )
    return rows


def _parse_xlsx(raw: bytes) -> list[dict]:
    try:
        from openpyxl import load_workbook
    except Exception:
        logger.exception("parse_xlsx: openpyxl unavailable")
        raise HTTPException(
            status_code=400,
            detail="XLSX import is unavailable, please upload a CSV file.",
        )
    try:
        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        header = next(rows_iter, None)
        if not header:
            return []
        keys = [str(h or "").strip().lower() for h in header]
        rows: list[dict] = []
        for values in rows_iter:
            if values is None or all(v in (None, "") for v in values):
                continue
            row: dict = {}
            for idx, key in enumerate(keys):
                if not key:
                    continue
                row[key] = values[idx] if idx < len(values) else ""
            rows.append(row)
        return rows
    except HTTPException:
        # Already a user-facing error with a concise detail, re-raise as is.
        logging.exception("Unexpected error")
        raise
    except Exception:
        logger.exception("parse_xlsx failed")
        raise HTTPException(
            status_code=400, detail="Could not read the uploaded spreadsheet."
        )


def row_label(index: int, row: dict, keys: tuple[str, ...]) -> str:
    """Human readable reference for a spreadsheet row."""
    tag = ""
    for key in keys:
        value = str((row or {}).get(key) or "").strip()
        if value:
            tag = value
            break
    return f"Row {index} [{tag or '(unnamed)'}]"


def format_validation_error(exc: ValidationError) -> str:
    """Flatten a pydantic error into one concise sentence."""
    parts: list[str] = []
    for err in exc.errors():
        loc = ".".join(str(x) for x in err.get("loc", ()) if x != "__root__")
        msg = str(err.get("msg", "invalid value")).replace("Value error, ", "")
        parts.append(f"{loc}: {msg}" if loc else msg)
    return "; ".join(parts) or "invalid row"


def normalize_row(raw_row: dict, fields: tuple[str, ...]) -> dict:
    """Keep only known, non-blank columns so schema defaults still apply."""
    row = {str(k or "").strip().lower(): v for k, v in (raw_row or {}).items()}
    payload = {f: row.get(f) for f in fields if f in row}
    return {
        k: (v.strip() if isinstance(v, str) else v)
        for k, v in payload.items()
        if v is not None and str(v).strip() != ""
    }
